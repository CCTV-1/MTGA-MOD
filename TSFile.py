import collections
import csv
import enum
import json


class TSFileType(enum.Enum):
    XLSX = 0,
    JSON = 1,
    CSV = 2


def loadTSInfo(filePath: str, type: TSFileType) -> collections.OrderedDict:
    TSInfo = collections.OrderedDict()
    match type:
        case TSFileType.JSON:
            with open(filePath, 'r', encoding='UTF-8') as f:
                TSInfo.update(json.load(f))
            return TSInfo
        case TSFileType.XLSX:
            import openpyxl
            TSBook = openpyxl.load_workbook(filePath, read_only=True)
            sheet = TSBook.active
            currentRow = 2
            while currentRow <= sheet.max_row:
                tsKey = sheet.cell(row=currentRow, column=1).value
                tsOracleText = sheet.cell(row=currentRow, column=2).value
                tsTranslation = sheet.cell(row=currentRow, column=3).value
                TSInfo[tsKey] = {'oracleText': tsOracleText,
                                'translation': tsTranslation}
                currentRow += 1
            return TSInfo
        case TSFileType.CSV:
            with open(filePath, 'r', encoding='UTF-8') as f:
                csvReader = csv.reader(f, delimiter=',', quotechar='"')
                for row in csvReader:
                    if len(row) != 3:
                        continue
                    TSInfo[row[0]] = {'oracleText': row[1], 'translation': row[2]}
            return TSInfo


def SaveTSInfo(TSInfo: collections.OrderedDict, name: str, type: TSFileType):
    match type:
        case TSFileType.XLSX:
            import openpyxl
            wb = openpyxl.Workbook()
            sheet = wb.create_sheet(name, 0)
            sheet.cell(1, 1, "键")
            sheet.cell(1, 2, "原文")
            sheet.cell(1, 3, "翻译")
            currentRow = 2
            for key, info in TSInfo.items():
                sheet.cell(currentRow, 1, key)
                sheet.cell(currentRow, 2, info['oracleText'])
                sheet.cell(currentRow, 3, info['translation'])
                currentRow += 1
            wb.save('{0}.xlsx'.format(name))
        case TSFileType.JSON:
            with open('{0}.json'.format(name), 'w', encoding='UTF-8') as f:
                json.dump(TSInfo, f, ensure_ascii=False, indent=4)
        case TSFileType.CSV:
            with open('{0}.csv'.format(name), 'w', encoding='UTF8', newline='') as f:
                fieldnames = ['key', 'oracle', 'ts']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                for key, value in TSInfo.items():
                    if value['oracleText'] == value['translation']:
                        writer.writerow({'key': key, 'oracle': value['oracleText'], 'ts': ""})
                    else:
                        writer.writerow({'key': key, 'oracle': value['oracleText'], 'ts': value['translation']})